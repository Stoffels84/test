# app.py
import re
import json
import base64
import html
import datetime as dt
from io import BytesIO
from urllib.parse import quote

import pandas as pd
import streamlit as st
import openpyxl
import bcrypt
import requests
import time

from pathlib import Path

# ----------------------------
# App local files (CSS/Logo/JSON remain local)
# ----------------------------
APP_DIR = Path(__file__).parent
CSS_PATH = APP_DIR / "styles.css"
LOGO_PATH = APP_DIR / "logo.png"
PERSONEEL_JSON_PATH = APP_DIR / "personeelsficheGB.json"

# ----------------------------
# Remote data config (Excel files)
# ----------------------------
DATA_BASE_URL = st.secrets.get("DATA_BASE_URL", "https://otgent.borolo.be/data").rstrip("/")
HOST_USER = st.secrets.get("HOST_USER", "")
HOST_PASS = st.secrets.get("HOST_PASS", "")

def data_url(filename: str) -> str:
    # encode filename safely (spaces, parentheses, etc.)
    return f"{DATA_BASE_URL}/{quote(filename)}"

@st.cache_data(show_spinner=False)
def fetch_bytes(url: str) -> bytes:
    if not HOST_USER or not HOST_PASS:
        raise ValueError("HOST_USER/HOST_PASS ontbreken in Streamlit secrets.")

    r = requests.get(url, auth=(HOST_USER, HOST_PASS), timeout=30)
    r.raise_for_status()
    return r.content

# Remote filenames
TOEGESTAAN_XLSX_NAME = "toegestaan_gebruik.xlsx"
XLSM_NAME = "schade met macro.xlsm"
GESPREKKEN_XLSX_NAME = "Overzicht gesprekken (aangepast).xlsx"
COACHINGS_XLSX_NAME = "Coachingslijst.xlsx"

# ----------------------------
# Sheets / Config
# ----------------------------
SCHADESHEET = "BRON"
GESPREKKEN_SHEET_NAME = "gesprekken per thema"
COACHINGS_SHEET_VOLTOOID = "Voltooide coachings"
COACHINGS_SHEET_COACHING = "Coaching"

# BRON columns to load (including teamcoach from BRON)
SCHADE_COLS = [
    "personeelsnr",
    "volledige naam",
    "teamcoach",
    "Datum",
    "Link",
    "Locatie",
    "voertuig",
    "bus/tram",
    "type",
]

PAGES = [
    ("dashboard", "Dashboard"),
    ("chauffeur", "Chauffeur"),
    ("voertuig", "Voertuig"),
    ("locatie", "Locatie"),
    ("coaching", "Coaching"),
    ("analyse", "Analyse"),
]

# ----------------------------
# Helpers
# ----------------------------
def set_progress(bar, text_ph, current, total, label):
    pct = int(current / total * 100)
    bar.progress(pct)
    text_ph.info(f"⏳ Bezig met laden: {label} ({current}/{total})")





def load_css(path: Path) -> None:
    """Load CSS from external file and inject into Streamlit."""
    if not path.exists():
        st.warning(f"CSS-bestand niet gevonden: {path.name} (zet dit naast app.py)")
        return
    css = path.read_text(encoding="utf-8")
    st.markdown(f"<style>{css}</style>", unsafe_allow_html=True)

def norm(s) -> str:
    return str(s).strip().lower()

def clean_id(v) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    if not s:
        return ""
    if re.fullmatch(r"\d+\.0", s):
        s = s[:-2]
    return s.strip()

def clean_text(v) -> str:
    return "" if v is None else str(v).strip()

def parse_year(v) -> int | None:
    if v is None:
        return None

    # Excel serial date (soms als getal)
    if isinstance(v, (int, float)) and 30000 < float(v) < 60000:
        try:
            base = dt.datetime(1899, 12, 30)
            d = base + dt.timedelta(days=float(v))
            return d.year
        except Exception:
            pass

    if isinstance(v, (dt.date, dt.datetime)):
        return v.year

    s = str(v).strip()
    if not s:
        return None

    m = re.match(r"^(\d{1,2})[/-](\d{1,2})[/-](\d{4})", s)
    if m:
        return int(m.group(3))

    m2 = re.match(r"^(\d{4})[/-](\d{1,2})[/-](\d{1,2})", s)
    if m2:
        return int(m2.group(1))

    try:
        return dt.datetime.fromisoformat(s).year
    except Exception:
        return None

def format_ddmmyyyy(v) -> str:
    """Toon altijd dd-mm-jjjj; tijd/uurnotatie verdwijnt."""
    if v is None:
        return ""
    s = str(v).strip()
    if not s:
        return ""
    try:
        ts = pd.to_datetime(s, dayfirst=True, errors="coerce")
        if pd.isna(ts):
            return s
        return ts.strftime("%d-%m-%Y")
    except Exception:
        return s

def img_to_data_uri(path: Path) -> str:
    b = path.read_bytes()
    ext = path.suffix.lower().lstrip(".")
    mime = "png" if ext == "png" else ext
    return f"data:image/{mime};base64,{base64.b64encode(b).decode('utf-8')}"

def _find_col(df: pd.DataFrame, wanted: str) -> str | None:
    w = norm(wanted)

    for c in df.columns:
        if norm(c) == w:
            return c

    if w in ["nummer", "personeelsnr", "personeelsnummer", "p-nr", "p_nr", "p nr", "p-nr."]:
        for alt in [
            "nr",
            "id",
            "persnr",
            "personeelsnr",
            "personeelsnummer",
            "nummer",
            "employeeid",
            "employee_id",
            "p-nr",
            "p nr",
            "p_nr",
            "p-nr.",
            "p-nr (p-nr)",
        ]:
            for c in df.columns:
                if norm(c) == norm(alt):
                    return c

    if w == "datum":
        for alt in ["date", "datum gesprek", "gespreksdatum", "datum coaching", "coachingsdatum"]:
            for c in df.columns:
                if norm(c) == norm(alt):
                    return c

    if w == "info":
        for alt in [
            "informatie",
            "opmerking",
            "opmerkingen",
            "beschrijving",
            "details",
            "thema",
            "onderwerp",
            "samenvatting",
            "actiepunten",
            "resultaat",
            "notities",
            "commentaar",
            "opmerkingen (coach)",
            "opmerkingen chauffeur",
            "opmerkingen",
        ]:
            for c in df.columns:
                if norm(c) == norm(alt):
                    return c

    if w in ["volledige naam", "chauffeurnaam", "naam"]:
        for alt in [
            "chauffeurnaam",
            "chauffeur naam",
            "naam",
            "medewerker",
            "werknemer",
            "chauffeur",
            "volledige naam",
            "full name",
            "fullname",
            "displayname",
            "display_name",
        ]:
            for c in df.columns:
                if norm(c) == norm(alt):
                    return c

    return None

def _flatten_json_to_records(data):
    if data is None:
        return []
    if isinstance(data, list):
        return [x for x in data if isinstance(x, dict)]
    if isinstance(data, dict):
        for k in ["data", "items", "results", "records"]:
            if k in data:
                return _flatten_json_to_records(data[k])
        if data and all(isinstance(v, dict) for v in data.values()):
            out = []
            for key, val in data.items():
                rec = dict(val)
                rec["_key"] = str(key)
                out.append(rec)
            return out
        return [data]
    return []

def render_html_table(
    df: pd.DataFrame,
    col_order: list[str],
    col_widths: dict[str, str],
    max_height_px: int = 520,
) -> None:
    view = df[col_order].copy()
    for c in col_order:
        view[c] = view[c].fillna("").astype(str)

    ths = []
    for c in col_order:
        w = col_widths.get(c, "auto")
        ths.append(f'<th style="width:{w}">{html.escape(c)}</th>')
    thead = "<tr>" + "".join(ths) + "</tr>"

    trs = []
    for _, row in view.iterrows():
        tds = []
        for c in col_order:
            cell = row[c]
            safe = html.escape(cell).replace("\n", "<br/>")
            tds.append(f"<td>{safe}</td>")
        trs.append("<tr>" + "".join(tds) + "</tr>")
    tbody = "".join(trs)

    st.markdown(
        f"""
        <div class="ot-table-wrap" style="max-height:{max_height_px}px;">
          <table class="ot-table">
            <thead>{thead}</thead>
            <tbody>{tbody}</tbody>
          </table>
        </div>
        """,
        unsafe_allow_html=True,
    )

# ----------------------------
# Login / Users (REMOTE toegestaan_gebruik.xlsx)
# ----------------------------
@st.cache_data(show_spinner=False)
def load_users_df() -> pd.DataFrame:
    url = data_url(TOEGESTAAN_XLSX_NAME)
    content = fetch_bytes(url)

    df = pd.read_excel(BytesIO(content), dtype=str).fillna("")
    df.columns = [c.strip().lower() for c in df.columns]

    # Verwacht: naam, rol, paswoord_hash (aanrader)
    if "naam" not in df.columns or "rol" not in df.columns:
        raise ValueError("Kolommen 'naam' en 'rol' zijn verplicht in toegestaan_gebruik.xlsx")

    # Ondersteun beide: paswoord_hash (aanrader) of paswoord (fallback)
    if "paswoord_hash" not in df.columns and "paswoord" not in df.columns:
        raise ValueError("Voor login heb je 'paswoord_hash' (aanrader) of 'paswoord' nodig.")

    df["naam"] = df["naam"].astype(str).str.strip()
    df["rol"] = df["rol"].astype(str).str.strip().str.lower()

    if "paswoord_hash" in df.columns:
        df["paswoord_hash"] = df["paswoord_hash"].astype(str).str.strip()
    if "paswoord" in df.columns:
        df["paswoord"] = df["paswoord"].astype(str).str.strip()

    # Uniek per naam
    df = df[df["naam"] != ""].copy()
    df = df.drop_duplicates(subset=["naam"], keep="last")

    return df

def verify_password(entered: str, row: pd.Series) -> bool:
    entered = (entered or "").strip()
    if not entered:
        return False

    # Aanrader: bcrypt hash check
    if "paswoord_hash" in row and str(row["paswoord_hash"]).strip():
        try:
            return bcrypt.checkpw(entered.encode("utf-8"), row["paswoord_hash"].encode("utf-8"))
        except Exception:
            return False

    # Fallback (niet ideaal): plain text vergelijken
    if "paswoord" in row and str(row["paswoord"]).strip():
        return entered == str(row["paswoord"]).strip()

    return False

def require_login() -> None:
    if st.session_state.get("auth_ok"):
        return

    st.title("🔐 Inloggen OT Gent")
    st.caption("Toegang is beveiligd. Meld aan om verder te gaan.")

    try:
        with st.spinner("Gebruikerslijst ophalen…"):
            users = load_users_df()
    except Exception as e:
        st.error("Kan toegestaan_gebruik.xlsx niet ophalen/lezen van de server.")
        st.exception(e)
        st.stop()

    naam = st.text_input("Naam", placeholder="bv. janssens", key="login_naam")
    pw = st.text_input("Paswoord", type="password", key="login_pw")

    c1, c2 = st.columns([1, 2])
    with c1:
        do_login = st.button("Inloggen", use_container_width=True)

    if do_login:
        naam_clean = (naam or "").strip()
        match = users[users["naam"] == naam_clean]

        if match.empty:
            st.error("Onbekende gebruiker.")
            st.stop()

        row = match.iloc[0]
        if verify_password(pw, row):
            st.session_state["auth_ok"] = True
            st.session_state["user_naam"] = row["naam"]
            st.session_state["user_rol"] = row.get("rol", "viewer")
            st.success("Ingelogd.")
            st.rerun()
        else:
            st.error("Onjuist paswoord.")
            st.stop()

    st.stop()


def logout_button() -> None:
    with st.sidebar:
        st.markdown("---")
        st.write(f"👤 **{st.session_state.get('user_naam','')}**")
        st.write(f"🔑 Rol: **{st.session_state.get('user_rol','')}**")
        if st.button("Uitloggen"):
            for k in ["auth_ok", "user_naam", "user_rol"]:
                st.session_state.pop(k, None)
            st.rerun()

# ----------------------------
# Navigation state
# ----------------------------
def get_page(default="dashboard") -> str:
    try:
        v = st.query_params.get("page", default)
        if isinstance(v, list):
            v = v[0] if v else default
        v = str(v).strip().lower()
    except Exception:
        v = default
    valid = {pid for pid, _ in PAGES}
    return v if v in valid else default

def set_page(page_id: str) -> None:
    st.query_params["page"] = page_id
    st.rerun()

# ----------------------------
# Remote Excel Loaders
# ----------------------------
@st.cache_data(show_spinner=False)
def load_schade_df() -> pd.DataFrame:
    url = data_url(XLSM_NAME)
    content = fetch_bytes(url)

    wb = openpyxl.load_workbook(BytesIO(content), data_only=True, keep_vba=True)
    if SCHADESHEET not in wb.sheetnames:
        raise ValueError(f"Tabblad '{SCHADESHEET}' niet gevonden in {XLSM_NAME}")

    ws = wb[SCHADESHEET]

    header = [c.value for c in ws[1]]
    header_map = {}
    for idx, h in enumerate(header):
        if h is None:
            continue
        key = norm(h)
        if key and key not in header_map:
            header_map[key] = idx

    def find_idx(col: str) -> int | None:
        key = norm(col)
        if key in header_map:
            return header_map[key]

        if col == "bus/tram":
            for alt in ["bus/ tram", "bus / tram", "bus - tram"]:
                if alt in header_map:
                    return header_map[alt]
        if col == "volledige naam":
            for alt in ["naam", "volledige naam.", "volledige naam "]:
                if alt in header_map:
                    return header_map[alt]
        if col == "teamcoach":
            for alt in ["team coach", "team_coach", "coach", "teamcoach "]:
                if alt in header_map:
                    return header_map[alt]

        if col == "voertuig":
            for c in header_map:
                if "voertuig" in c.replace(" ", ""):
                    return header_map[c]

        return None

    idx_map = {c: find_idx(c) for c in SCHADE_COLS}

    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        obj = {}
        any_val = False

        for col in SCHADE_COLS:
            i = idx_map.get(col)
            val = r[i] if (i is not None and i < len(r)) else None

            if val is not None and str(val).strip() != "":
                any_val = True

            if col == "Datum" and isinstance(val, (dt.date, dt.datetime)):
                val = val.isoformat()

            obj[col] = val

        if any_val:
            rows.append(obj)

    df = pd.DataFrame(rows)
    for c in SCHADE_COLS:
        if c not in df.columns:
            df[c] = None

    df["personeelsnr"] = df["personeelsnr"].apply(clean_id)
    df["volledige naam"] = df["volledige naam"].apply(clean_text)
    df["teamcoach"] = df["teamcoach"].apply(clean_text)
    df["voertuig"] = df["voertuig"].apply(clean_text)

    df["_jaar"] = df["Datum"].apply(parse_year)
    df["_search"] = (
        df["personeelsnr"].fillna("").astype(str)
        + " "
        + df["volledige naam"].fillna("").astype(str)
        + " "
        + df["teamcoach"].fillna("").astype(str)
        + " "
        + df["voertuig"].fillna("").astype(str)
    ).str.lower()

    return df

@st.cache_data(show_spinner=False)
def load_gesprekken_df() -> pd.DataFrame:
    url = data_url(GESPREKKEN_XLSX_NAME)
    content = fetch_bytes(url)
    bio = BytesIO(content)

    xls = pd.ExcelFile(bio)
    if GESPREKKEN_SHEET_NAME not in xls.sheet_names:
        raise ValueError(
            f"Tabblad '{GESPREKKEN_SHEET_NAME}' niet gevonden in {GESPREKKEN_XLSX_NAME}. "
            f"Gevonden tabs: {xls.sheet_names}"
        )

    df = pd.read_excel(BytesIO(content), sheet_name=GESPREKKEN_SHEET_NAME, dtype=str)

    num_col = _find_col(df, "nummer")
    date_col = _find_col(df, "Datum")
    info_col = _find_col(df, "Info")
    name_col = _find_col(df, "Chauffeurnaam")

    if num_col is None:
        raise ValueError("Kolom 'nummer' (personeelsnr) niet gevonden in 'gesprekken per thema'.")

    if num_col != "nummer":
        df = df.rename(columns={num_col: "nummer"})
    if date_col and date_col != "Datum":
        df = df.rename(columns={date_col: "Datum"})
    if info_col and info_col != "Info":
        df = df.rename(columns={info_col: "Info"})
    if name_col and name_col != "Chauffeurnaam":
        df = df.rename(columns={name_col: "Chauffeurnaam"})

    for c in ["Datum", "Info", "Chauffeurnaam"]:
        if c not in df.columns:
            df[c] = ""

    df["nummer"] = df["nummer"].apply(clean_id)
    df["Datum"] = df["Datum"].apply(clean_text)
    df["Info"] = df["Info"].apply(clean_text)
    df["Chauffeurnaam"] = df["Chauffeurnaam"].apply(clean_text)

    df["_search"] = (
        df["nummer"].fillna("").astype(str)
        + " "
        + df["Chauffeurnaam"].fillna("").astype(str)
        + " "
        + df["Info"].fillna("").astype(str)
    ).str.lower()
    df["_jaar"] = df["Datum"].apply(parse_year)
    return df

@st.cache_data(show_spinner=False)
def fetch_coachings_bytes() -> bytes:
    return fetch_bytes(data_url(COACHINGS_XLSX_NAME))

@st.cache_data(show_spinner=False)
def load_coaching_voltooid_df() -> pd.DataFrame:
    content = fetch_coachings_bytes()
    xls = pd.ExcelFile(BytesIO(content))

    if COACHINGS_SHEET_VOLTOOID not in xls.sheet_names:
        raise ValueError(
            f"Tabblad '{COACHINGS_SHEET_VOLTOOID}' niet gevonden in {COACHINGS_XLSX_NAME}. "
            f"Gevonden tabs: {xls.sheet_names}"
        )

    df = pd.read_excel(BytesIO(content), sheet_name=COACHINGS_SHEET_VOLTOOID, dtype=str)

    num_col = _find_col(df, "nummer") or _find_col(df, "personeelsnr")
    name_col = _find_col(df, "Chauffeurnaam") or _find_col(df, "naam") or _find_col(df, "volledige naam")
    date_col = _find_col(df, "Datum")
    info_col = _find_col(df, "Info")

    if num_col is None:
        df["nummer"] = ""
    else:
        if num_col != "nummer":
            df = df.rename(columns={num_col: "nummer"})

    if name_col is None:
        df["Chauffeurnaam"] = ""
    else:
        if name_col != "Chauffeurnaam":
            df = df.rename(columns={name_col: "Chauffeurnaam"})

    if date_col is None:
        df["Datum"] = ""
    else:
        if date_col != "Datum":
            df = df.rename(columns={date_col: "Datum"})

    if info_col is None:
        candidates = []
        for c in df.columns:
            if norm(c) in [
                "thema",
                "onderwerp",
                "opmerking",
                "opmerkingen",
                "samenvatting",
                "notities",
                "commentaar",
                "actiepunten",
                "resultaat",
            ]:
                candidates.append(c)
        if candidates:
            df["Info"] = df[candidates].fillna("").astype(str).agg(" | ".join, axis=1)
        else:
            df["Info"] = ""
    else:
        if info_col != "Info":
            df = df.rename(columns={info_col: "Info"})

    df["nummer"] = df["nummer"].apply(clean_id)
    df["Chauffeurnaam"] = df["Chauffeurnaam"].apply(clean_text)
    df["Datum"] = df["Datum"].apply(clean_text)
    df["Info"] = df["Info"].apply(clean_text)

    df["_search"] = (
        df["nummer"].fillna("").astype(str)
        + " "
        + df["Chauffeurnaam"].fillna("").astype(str)
        + " "
        + df["Info"].fillna("").astype(str)
    ).str.lower()
    df["_jaar"] = df["Datum"].apply(parse_year)
    return df

@st.cache_data(show_spinner=False)
def load_coaching_tab_df() -> pd.DataFrame:
    """
    Coachingslijst.xlsx -> tab 'Coaching'
    Kolommen: P-nr, Volledige naam, Opmerkingen
    """
    content = fetch_coachings_bytes()
    xls = pd.ExcelFile(BytesIO(content))

    if COACHINGS_SHEET_COACHING not in xls.sheet_names:
        raise ValueError(
            f"Tabblad '{COACHINGS_SHEET_COACHING}' niet gevonden in {COACHINGS_XLSX_NAME}. "
            f"Gevonden tabs: {xls.sheet_names}"
        )

    df = pd.read_excel(BytesIO(content), sheet_name=COACHINGS_SHEET_COACHING, dtype=str)

    pnr_col = _find_col(df, "P-nr") or _find_col(df, "nummer") or _find_col(df, "personeelsnr")
    name_col = _find_col(df, "Volledige naam") or _find_col(df, "naam") or _find_col(df, "chauffeurnaam")
    opm_col = _find_col(df, "Opmerkingen") or _find_col(df, "Info")

    if pnr_col is None:
        df["nummer"] = ""
    else:
        if pnr_col != "nummer":
            df = df.rename(columns={pnr_col: "nummer"})

    if name_col is None:
        df["Chauffeurnaam"] = ""
    else:
        if name_col != "Chauffeurnaam":
            df = df.rename(columns={name_col: "Chauffeurnaam"})

    if opm_col is None:
        df["Info"] = ""
    else:
        if opm_col != "Info":
            df = df.rename(columns={opm_col: "Info"})

    df["nummer"] = df["nummer"].apply(clean_id)
    df["Chauffeurnaam"] = df["Chauffeurnaam"].apply(clean_text)
    df["Info"] = df["Info"].apply(clean_text)

    df["_search"] = (
        df["nummer"].fillna("").astype(str)
        + " "
        + df["Chauffeurnaam"].fillna("").astype(str)
        + " "
        + df["Info"].fillna("").astype(str)
    ).str.lower()
    return df

@st.cache_data(show_spinner=False)
def load_personeelsfiche_df() -> pd.DataFrame:
    if not PERSONEEL_JSON_PATH.exists():
        return pd.DataFrame(columns=["_search"])

    try:
        data = json.loads(PERSONEEL_JSON_PATH.read_text(encoding="utf-8"))
    except Exception:
        data = json.loads(PERSONEEL_JSON_PATH.read_text())

    records = _flatten_json_to_records(data)
    if not records:
        return pd.DataFrame(columns=["_search"])

    df = pd.DataFrame(records)

    id_col = _find_col(df, "personeelsnr") or _find_col(df, "nummer") or _find_col(df, "personeelsnummer")
    name_col = _find_col(df, "volledige naam") or _find_col(df, "naam") or _find_col(df, "chauffeurnaam")

    if id_col is None and "_key" in df.columns:
        id_col = "_key"

    if id_col and id_col != "personeelsnr":
        df = df.rename(columns={id_col: "personeelsnr"})
        id_col = "personeelsnr"
    if name_col and name_col != "naam":
        df = df.rename(columns={name_col: "naam"})
        name_col = "naam"

    if id_col is None:
        df["personeelsnr"] = ""
        id_col = "personeelsnr"
    if name_col is None:
        df["naam"] = ""
        name_col = "naam"

    df[id_col] = df[id_col].apply(clean_id)
    df[name_col] = df[name_col].apply(clean_text)

    extra_cols = []
    for c in df.columns:
        if c in ["_search", id_col, name_col]:
            continue
        if norm(c) in ["dienst", "afdeling", "team", "functie", "rol", "standplaats", "locatie", "teamcoach"]:
            extra_cols.append(c)

    parts = [df[id_col].fillna("").astype(str), df[name_col].fillna("").astype(str)]
    for c in extra_cols[:6]:
        parts.append(df[c].fillna("").astype(str))

    df["_search"] = parts[0]
    for s in parts[1:]:
        df["_search"] = df["_search"].astype(str) + " " + s.astype(str)
    df["_search"] = df["_search"].str.lower()
    return df

# ----------------------------
# Streamlit setup
# ----------------------------
st.set_page_config(page_title="Analyse en rapportering OT Gent", layout="wide")
load_css(CSS_PATH)

require_login()
logout_button()

# ----------------------------
# Load data
# -------------

load_ph = st.empty()
with load_ph.container():
    st.info("📦 Data wordt geladen...")

    bar = st.progress(0)
    text_ph = st.empty()

    total = 5
    step = 0

    try:
        step += 1
        set_progress(bar, text_ph, step, total, "Schade (schade met macro.xlsm)")
        df_schade = load_schade_df()

        step += 1
        set_progress(bar, text_ph, step, total, "Gesprekken (Overzicht gesprekken.xlsx)")
        df_gesprekken = load_gesprekken_df()

        step += 1
        set_progress(bar, text_ph, step, total, "Voltooide coachings (Coachingslijst.xlsx)")
        df_coach_voltooid = load_coaching_voltooid_df()

        step += 1
        set_progress(bar, text_ph, step, total, "Geplande coachings (Coachingslijst.xlsx)")
        df_coach_tab = load_coaching_tab_df()

        step += 1
        set_progress(bar, text_ph, step, total, "Personeelsfiche (JSON lokaal)")
        df_personeel = load_personeelsfiche_df()

        bar.progress(100)
        text_ph.success("🚀 Alle data succesvol geladen!")
        time.sleep(2)
        load_ph.empty()

    except Exception as e:
        text_ph.error("❌ Fout bij laden van data")
        st.exception(e)
        st.stop()


years_schade = df_schade["_jaar"].dropna().unique().tolist() if "_jaar" in df_schade.columns else []
years_gespr = df_gesprekken["_jaar"].dropna().unique().tolist() if "_jaar" in df_gesprekken.columns else []
years_volt = df_coach_voltooid["_jaar"].dropna().unique().tolist() if "_jaar" in df_coach_voltooid.columns else []
years = sorted({int(y) for y in (years_schade + years_gespr + years_volt) if y is not None}, reverse=True)

current_page = get_page("dashboard")

# ----------------------------
# Topbar
# ----------------------------
st.markdown('<div class="ot-topbar">', unsafe_allow_html=True)

c1, c2, c3 = st.columns([2.3, 1.2, 3.5], vertical_alignment="center")

with c1:
    logo_html = f'<img class="ot-logo" src="{img_to_data_uri(LOGO_PATH)}" alt="Logo" />' if LOGO_PATH.exists() else ""
    st.markdown(
        f"""
        <div class="ot-brand">
          {logo_html}
          <div>
            <div class="ot-title">Analyse en rapportering OT Gent</div>
            <div class="ot-sub">schade</div>
          </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

with c2:
    year_choice = st.selectbox("Jaar", ["Alle"] + [str(y) for y in years], index=0)

with c3:
    tab_cols = st.columns([1, 1, 1, 1, 1, 1, 0.95], gap="small")

    for (pid, label), col in zip(PAGES, tab_cols[:6]):
        with col:
            active = (pid == current_page)
            st.markdown(f'<div class="ot-tab-btn {"active" if active else ""}">', unsafe_allow_html=True)
            if st.button(label, key=f"tab_{pid}", use_container_width=True):
                set_page(pid)
            st.markdown("</div>", unsafe_allow_html=True)

    with tab_cols[6]:
        st.markdown('<div class="ot-tab-btn">', unsafe_allow_html=True)
        if st.button("↻ Herladen", key="reload_btn", use_container_width=True):
            st.cache_data.clear()
            try:
                st.cache_resource.clear()
            except Exception:
                pass
            st.rerun()
        st.markdown("</div>", unsafe_allow_html=True)

st.markdown("</div>", unsafe_allow_html=True)

# ----------------------------
# Year filter views
# ----------------------------
df_schade_view = df_schade[df_schade["_jaar"] == int(year_choice)].copy() if year_choice != "Alle" else df_schade.copy()
df_gesprekken_view = df_gesprekken[df_gesprekken["_jaar"] == int(year_choice)].copy() if year_choice != "Alle" else df_gesprekken.copy()
df_coach_voltooid_view = (
    df_coach_voltooid[df_coach_voltooid["_jaar"] == int(year_choice)].copy()
    if year_choice != "Alle"
    else df_coach_voltooid.copy()
)
# df_coach_tab heeft geen jaarfilter (geen datumkolom)

# ----------------------------
# Pages
# ----------------------------
if current_page == "dashboard":
    st.subheader("Dashboard (met update om 1u en 13u)")

    q = st.text_input(
        "Zoek op personeelsnr of naam.",
        placeholder="Typ om te zoeken…",
    ).strip().lower()

    if not q:
        st.caption("Typ iets in het zoekveld om resultaten te zien.")
        st.stop()

    schade_hits = df_schade_view[df_schade_view["_search"].str.contains(re.escape(q), na=False)].copy()
    gesprekken_hits = df_gesprekken_view[df_gesprekken_view["_search"].str.contains(re.escape(q), na=False)].copy()
    coach_volt_hits = df_coach_voltooid_view[df_coach_voltooid_view["_search"].str.contains(re.escape(q), na=False)].copy()

    coach_tab_hits = pd.DataFrame()
    if "_search" in df_coach_tab.columns and len(df_coach_tab) > 0:
        coach_tab_hits = df_coach_tab[df_coach_tab["_search"].str.contains(re.escape(q), na=False)].copy()

    personeels_hits = pd.DataFrame()
    if "_search" in df_personeel.columns and len(df_personeel) > 0:
        personeels_hits = df_personeel[df_personeel["_search"].str.contains(re.escape(q), na=False)].copy()

    # Personeelsfiche
    st.markdown("#### Personeelsfiche")
    if len(personeels_hits) == 0:
        st.caption("Geen personeelsfiche gevonden voor deze zoekterm.")
    else:
        summary_cols = [c for c in ["personeelsnr", "naam"] if c in personeels_hits.columns]
        if summary_cols:
            st.dataframe(personeels_hits[summary_cols].head(20), use_container_width=True, hide_index=True)

        max_show = 10
        for i, (_, row) in enumerate(personeels_hits.head(max_show).iterrows(), start=1):
            pid = row.get("personeelsnr", "")
            nm = row.get("naam", "")
            title = f"{i}. {pid} — {nm}".strip(" —")
            with st.expander(title, expanded=(i == 1)):
                rec = row.drop(labels=["_search"], errors="ignore").to_dict()
                st.json(rec)

        if len(personeels_hits) > max_show:
            st.caption(f"… en nog {len(personeels_hits) - max_show} extra matches.")

    # Schade
    st.markdown("#### Schade (BRON)")
    if len(schade_hits) == 0:
        st.caption("Geen schadegevallen gevonden voor deze zoekterm.")
    else:
        show_cols = [c for c in SCHADE_COLS if c in schade_hits.columns]
        show = schade_hits[show_cols].head(500).copy()
        if "Datum" in show.columns:
            show["Datum"] = show["Datum"].apply(format_ddmmyyyy)

        column_config = {
            "personeelsnr": st.column_config.TextColumn("personeelsnr", width="small"),
            "volledige naam": st.column_config.TextColumn("volledige naam", width="medium"),
            "teamcoach": st.column_config.TextColumn("teamcoach", width="medium"),
            "Datum": st.column_config.TextColumn("Datum", width="small"),
            "Link": st.column_config.LinkColumn("Open EAF", display_text="Open EAF", width="small"),
            "Locatie": st.column_config.TextColumn("Locatie", width="medium"),
            "voertuig": st.column_config.TextColumn("voertuig", width="medium"),
            "bus/tram": st.column_config.TextColumn("bus/tram", width="small"),
            "type": st.column_config.TextColumn("type", width="small"),
        }

        st.dataframe(
            show,
            use_container_width=True,
            hide_index=True,
            column_config=column_config,
        )

    # Geplande coaching
    st.markdown("#### Geplande coaching")
    if len(coach_tab_hits) == 0:
        st.caption("Geen geplande coaching-info gevonden voor deze zoekterm.")
    else:
        display_ct = coach_tab_hits[["nummer", "Chauffeurnaam", "Info"]].copy()
        render_html_table(
            display_ct.head(300),
            col_order=["nummer", "Chauffeurnaam", "Info"],
            col_widths={"nummer": "90px", "Chauffeurnaam": "220px", "Info": "auto"},
            max_height_px=520,
        )

    # Voltooide coaching
    st.markdown("#### Voltooide coaching")
    if len(coach_volt_hits) == 0:
        st.caption("Geen voltooide coachings gevonden voor deze zoekterm.")
    else:
        display_v = coach_volt_hits[["nummer", "Chauffeurnaam", "Datum", "Info"]].copy()
        display_v["Datum"] = display_v["Datum"].apply(format_ddmmyyyy)
        render_html_table(
            display_v.head(300),
            col_order=["nummer", "Chauffeurnaam", "Datum", "Info"],
            col_widths={"nummer": "90px", "Chauffeurnaam": "180px", "Datum": "120px", "Info": "auto"},
            max_height_px=520,
        )

    # Overzicht gesprekken
    st.markdown("#### Overzicht gesprekken")
    if len(gesprekken_hits) == 0:
        st.caption("Geen gesprekken gevonden voor deze zoekterm.")
    else:
        display_g = gesprekken_hits[["nummer", "Chauffeurnaam", "Datum", "Info"]].copy()
        display_g["Datum"] = display_g["Datum"].apply(format_ddmmyyyy)
        render_html_table(
            display_g.head(300),
            col_order=["nummer", "Chauffeurnaam", "Datum", "Info"],
            col_widths={"nummer": "90px", "Chauffeurnaam": "180px", "Datum": "120px", "Info": "auto"},
            max_height_px=520,
        )


    # ----------------------------
    # Tijdlijn (onderaan dashboard)
    # ----------------------------
    st.divider()
    st.markdown("## 🧾 Tijdlijn")

    def _to_dt(v):
        return pd.to_datetime(v, dayfirst=True, errors="coerce")

    timeline_rows = []

    # 1) Schade
    if not schade_hits.empty:
        s = schade_hits.copy()
        s["_dt"] = s["Datum"].apply(_to_dt)
        for _, r in s.iterrows():
            summary_parts = []
            loc = str(r.get("Locatie", "") or "").strip()
            typ = str(r.get("type", "") or "").strip()
            veh = str(r.get("voertuig", "") or "").strip()
            bt  = str(r.get("bus/tram", "") or "").strip()
            tc  = str(r.get("teamcoach", "") or "").strip()

            if typ: summary_parts.append(f"Type: {typ}")
            if loc: summary_parts.append(f"Locatie: {loc}")
            if veh: summary_parts.append(f"Voertuig: {veh}")
            if bt:  summary_parts.append(f"Bus/Tram: {bt}")
            if tc:  summary_parts.append(f"Teamcoach: {tc}")

            timeline_rows.append({
                "Datum": r.get("Datum", ""),
                "_dt": r.get("_dt", pd.NaT),
                "Bron": "Schade",
                "P-nr": r.get("personeelsnr", ""),
                "Naam": r.get("volledige naam", ""),
                "Samenvatting": " | ".join(summary_parts) if summary_parts else "",
                "Link": r.get("Link", ""),
            })

    # 2) Overzicht gesprekken
    if not gesprekken_hits.empty:
        g = gesprekken_hits.copy()
        g["_dt"] = g["Datum"].apply(_to_dt)
        for _, r in g.iterrows():
            timeline_rows.append({
                "Datum": r.get("Datum", ""),
                "_dt": r.get("_dt", pd.NaT),
                "Bron": "Gesprek",
                "P-nr": r.get("nummer", ""),
                "Naam": r.get("Chauffeurnaam", ""),
                "Samenvatting": str(r.get("Info", "") or "").strip(),
                "Link": "",  # gesprekken hebben geen Link-kolom
            })

    # 3) Voltooide coaching
    if not coach_volt_hits.empty:
        v = coach_volt_hits.copy()
        v["_dt"] = v["Datum"].apply(_to_dt)
        for _, r in v.iterrows():
            timeline_rows.append({
                "Datum": r.get("Datum", ""),
                "_dt": r.get("_dt", pd.NaT),
                "Bron": "Coaching (voltooid)",
                "P-nr": r.get("nummer", ""),
                "Naam": r.get("Chauffeurnaam", ""),
                "Samenvatting": str(r.get("Info", "") or "").strip(),
                "Link": "",
            })

    # 4) Geplande coaching (geen datum aanwezig → NaT)
    if not coach_tab_hits.empty:
        ct = coach_tab_hits.copy()
        for _, r in ct.iterrows():
            timeline_rows.append({
                "Datum": "",              # geen datum
                "_dt": pd.NaT,            # zodat dit onderaan sorteert
                "Bron": "Coaching (gepland)",
                "P-nr": r.get("nummer", ""),
                "Naam": r.get("Chauffeurnaam", ""),
                "Samenvatting": str(r.get("Info", "") or "").strip(),
                "Link": "",
            })

    if not timeline_rows:
        st.caption("Geen items voor tijdlijn gevonden bij deze zoekterm.")
    else:
        tl = pd.DataFrame(timeline_rows)

        # sorteer: meest recent eerst, NaT onderaan
        tl = tl.sort_values(by="_dt", ascending=False, na_position="last").drop(columns=["_dt"])

        # Datum formateren (dd-mm-jjjj)
        tl["Datum"] = tl["Datum"].apply(format_ddmmyyyy)
        tl["Link"] = tl["Link"].replace({"": None})

        # Beperk aantal rijen (veilig voor performance)
        tl = tl.head(300)

        column_config = {
            "Datum": st.column_config.TextColumn("Datum", width="small"),
            "Bron": st.column_config.TextColumn("Bron", width="small"),
            "P-nr": st.column_config.TextColumn("P-nr", width="small"),
            "Naam": st.column_config.TextColumn("Naam", width="medium"),
            "Samenvatting": st.column_config.TextColumn("Samenvatting", width="large"),
        }
        if "Link" in tl.columns:
            column_config["Link"] = st.column_config.LinkColumn("Open EAF", display_text="Open EAF", width="small")

        st.dataframe(
            tl,
            use_container_width=True,
            hide_index=True,
            column_config=column_config,
        )

        st.caption("Tip: gebruik de zoekbalk bovenaan om de tijdlijn per chauffeur/personeelsnr te bekijken.")












elif current_page == "chauffeur":
    st.subheader("Chauffeur")

    if df_schade_view.empty:
        st.info("Geen schadegegevens beschikbaar voor deze selectie.")
        st.stop()

    # Controls
    top_n = st.selectbox("Top", [10, 20, 50, 100], index=1)
    min_aantal = st.slider("Minimum aantal schadegevallen", 1, 20, 1)

    # ---- Top chauffeurs ----
    st.markdown("### 🚗 Chauffeurs met meeste schadegevallen")

    top_chauffeurs = (
        df_schade_view
        .groupby(["personeelsnr", "volledige naam"], dropna=False)
        .size()
        .reset_index(name="Aantal schadegevallen")
        .sort_values("Aantal schadegevallen", ascending=False)
    )

    top_chauffeurs_filtered = top_chauffeurs[top_chauffeurs["Aantal schadegevallen"] >= min_aantal].head(top_n)

    st.dataframe(
        top_chauffeurs_filtered,
        use_container_width=True,
        hide_index=True,
        column_config={
            "personeelsnr": st.column_config.TextColumn("Personeelsnr", width="small"),
            "volledige naam": st.column_config.TextColumn("Chauffeur", width="medium"),
            "Aantal schadegevallen": st.column_config.NumberColumn("Aantal", width="small"),
        },
    )

    if len(top_chauffeurs_filtered) == 0:
        st.caption("Geen chauffeurs binnen deze filters.")

    # ---- Teamcoach ----
    st.markdown("### 👥 Teamcoach: aantal schadegevallen")

    if "teamcoach" not in df_schade_view.columns:
        st.warning("Kolom 'teamcoach' niet gevonden in BRON.")
        st.stop()

    schade_per_teamcoach = (
        df_schade_view
        .assign(teamcoach=df_schade_view["teamcoach"].fillna("").astype(str).str.strip())
        .replace({"teamcoach": {"": "(onbekend)"}})
        .groupby("teamcoach", dropna=False)
        .size()
        .reset_index(name="Aantal schadegevallen")
        .sort_values("Aantal schadegevallen", ascending=False)
    )

    st.dataframe(
        schade_per_teamcoach.rename(columns={"Aantal schadegevallen": "Aantal"}),
        use_container_width=True,
        hide_index=True,
        column_config={
            "teamcoach": st.column_config.TextColumn("Teamcoach", width="medium"),
            "Aantal": st.column_config.NumberColumn("Aantal", width="small"),
        },
    )

    schade_per_teamcoach_sorted = (
        schade_per_teamcoach
        .sort_values("Aantal schadegevallen", ascending=False)
        .set_index("teamcoach")
    )

    st.bar_chart(schade_per_teamcoach_sorted["Aantal schadegevallen"])

elif current_page == "voertuig":
    st.subheader("Voertuig")

    if df_schade_view.empty:
        st.info("Geen schadegegevens beschikbaar voor deze selectie.")
        st.stop()

    def _to_month(v) -> str:
        if v is None:
            return ""
        s = str(v).strip()
        if not s:
            return ""
        ts = pd.to_datetime(s, dayfirst=True, errors="coerce")
        if pd.isna(ts):
            return ""
        return ts.strftime("%Y-%m")

    c1, c2, c3, c4 = st.columns([1.0, 1.0, 1.1, 1.4])
    with c1:
        top_n = st.selectbox("Top", [10, 20, 50, 100, 200], index=1)
    with c2:
        min_aantal = st.slider("Minimum aantal schadegevallen", 1, 50, 1)
    with c3:
        bt_vals = (
            df_schade_view["bus/tram"]
            .fillna("")
            .astype(str)
            .str.strip()
            .replace("", "(onbekend)")
            .unique()
            .tolist()
        )
        bt_vals = sorted(bt_vals)
        bus_tram = st.selectbox("Bus/Tram", ["Alles"] + bt_vals, index=0)
    with c4:
        voertuig_q = st.text_input("Zoek voertuig", placeholder="bv. 6301, 7205, ...").strip().lower()

    tmp = df_schade_view.copy()

    tmp["voertuig"] = tmp["voertuig"].fillna("").astype(str).str.strip()
    tmp["bus/tram"] = tmp["bus/tram"].fillna("").astype(str).str.strip().replace("", "(onbekend)")
    tmp["Locatie"] = tmp["Locatie"].fillna("").astype(str).str.strip()
    tmp["type"] = tmp["type"].fillna("").astype(str).str.strip()
    tmp["teamcoach"] = tmp["teamcoach"].fillna("").astype(str).str.strip()

    if bus_tram != "Alles":
        tmp = tmp[tmp["bus/tram"] == bus_tram].copy()

    if voertuig_q:
        tmp = tmp[tmp["voertuig"].str.lower().str.contains(re.escape(voertuig_q), na=False)].copy()

    tmp["voertuig"] = tmp["voertuig"].replace("", "(onbekend)")

    total_cases = len(tmp)
    unique_voertuigen = tmp["voertuig"].nunique(dropna=True)
    avg_per_voertuig = (total_cases / unique_voertuigen) if unique_voertuigen else 0.0

    top_voertuig = ""
    top_voertuig_count = 0
    if total_cases > 0:
        vc = tmp.groupby("voertuig").size().sort_values(ascending=False)
        if len(vc) > 0:
            top_voertuig = str(vc.index[0])
            top_voertuig_count = int(vc.iloc[0])

    k1, k2, k3, k4 = st.columns(4)
    k1.metric("Schadegevallen", f"{total_cases}")
    k2.metric("Unieke voertuigen", f"{unique_voertuigen}")
    k3.metric("Gemiddeld / voertuig", f"{avg_per_voertuig:.2f}")
    k4.metric("Top voertuig", f"{top_voertuig_count} — {top_voertuig}" if top_voertuig else "—")

    st.divider()

    st.markdown("### 🚋 Top voertuigen met meeste schadegevallen")

    voertuigen_counts = (
        tmp.groupby("voertuig")
        .size()
        .reset_index(name="Aantal")
        .sort_values("Aantal", ascending=False)
    )

    voertuigen_counts = voertuigen_counts[voertuigen_counts["Aantal"] >= min_aantal].copy()

    def _mode_or_empty(s: pd.Series) -> str:
        s = s.dropna().astype(str).str.strip()
        s = s[s != ""]
        if s.empty:
            return ""
        return s.value_counts().index[0]

    extra = (
        tmp.groupby("voertuig", dropna=False)
        .agg(
            BusTram=("bus/tram", _mode_or_empty),
            LaatsteDatum=("Datum", lambda x: pd.to_datetime(x, dayfirst=True, errors="coerce").max()),
            TopLocatie=("Locatie", _mode_or_empty),
        )
        .reset_index()
    )

    extra["LaatsteDatum"] = pd.to_datetime(extra["LaatsteDatum"], errors="coerce")
    extra["LaatsteDatum"] = extra["LaatsteDatum"].dt.strftime("%d-%m-%Y").fillna("")

    top_table = voertuigen_counts.merge(extra, on="voertuig", how="left").head(top_n)

    st.dataframe(
        top_table,
        use_container_width=True,
        hide_index=True,
        column_config={
            "voertuig": st.column_config.TextColumn("Voertuig", width="medium"),
            "Aantal": st.column_config.NumberColumn("Aantal", width="small"),
            "BusTram": st.column_config.TextColumn("Bus/Tram (meest voork.)", width="small"),
            "LaatsteDatum": st.column_config.TextColumn("Laatste datum", width="small"),
            "TopLocatie": st.column_config.TextColumn("Top locatie", width="medium"),
        },
    )

    if top_table.empty:
        st.caption("Geen voertuigen binnen deze filters.")
        st.stop()

    st.divider()

    st.markdown("### 📈 Trend & details voor gekozen voertuig")

    voertuig_options = top_table["voertuig"].tolist()
    gekozen_voertuig = st.selectbox("Kies voertuig", voertuig_options, index=0)

    vdf = tmp[tmp["voertuig"] == gekozen_voertuig].copy()

    vdf["Maand"] = vdf["Datum"].apply(_to_month)
    per_maand = (
        vdf[vdf["Maand"] != ""]
        .groupby("Maand")
        .size()
        .reset_index(name="Aantal")
        .sort_values("Maand")
    )

    cL, cR = st.columns([1.2, 1.0], gap="large")
    with cL:
        st.markdown("#### Schade per maand")
        if per_maand.empty:
            st.caption("Geen geldige datums om per maand te groeperen.")
        else:
            st.dataframe(per_maand, use_container_width=True, hide_index=True)
            st.bar_chart(per_maand.set_index("Maand")["Aantal"])

    with cR:
        st.markdown("#### Breakdown (top 10)")
        per_type = (
            vdf.assign(type=vdf["type"].replace("", "(onbekend)"))
            .groupby("type")
            .size()
            .reset_index(name="Aantal")
            .sort_values("Aantal", ascending=False)
            .head(10)
        )
        st.caption("Type")
        st.dataframe(per_type, use_container_width=True, hide_index=True)

        per_loc = (
            vdf.assign(Locatie=vdf["Locatie"].replace("", "(onbekend)"))
            .groupby("Locatie")
            .size()
            .reset_index(name="Aantal")
            .sort_values("Aantal", ascending=False)
            .head(10)
        )
        st.caption("Locatie")
        st.dataframe(per_loc, use_container_width=True, hide_index=True)

    st.markdown("#### Detail-lijst (laatste 200)")
    detail_cols = [c for c in ["Datum", "Locatie", "type", "bus/tram", "teamcoach", "volledige naam", "personeelsnr", "Link"] if c in vdf.columns]
    details = vdf[detail_cols].copy()

    if "Datum" in details.columns:
        details["Datum"] = details["Datum"].apply(format_ddmmyyyy)

    try:
        sort_ts = pd.to_datetime(vdf["Datum"].astype(str), dayfirst=True, errors="coerce")
        details["_sort"] = sort_ts
        details = details.sort_values("_sort", ascending=False).drop(columns=["_sort"])
    except Exception:
        pass

    if "Link" in details.columns:
        details["Link"] = details["Link"].replace({"": None})

    column_config = {}
    if "Link" in details.columns:
        column_config["Link"] = st.column_config.LinkColumn("Open EAF", display_text="Open EAF", width="small")

    st.dataframe(
        details.head(200),
        use_container_width=True,
        hide_index=True,
        column_config=column_config if column_config else None,
    )

elif current_page == "locatie":
    st.subheader("Locatie")

    if df_schade_view.empty:
        st.info("Geen schadegegevens beschikbaar voor deze selectie.")
        st.stop()

    def _to_month(v) -> str:
        if v is None:
            return ""
        s = str(v).strip()
        if not s:
            return ""
        ts = pd.to_datetime(s, dayfirst=True, errors="coerce")
        if pd.isna(ts):
            return ""
        return ts.strftime("%Y-%m")

    c1, c2, c3, c4, c5 = st.columns([1.0, 1.1, 1.1, 1.2, 1.4])
    with c1:
        top_n = st.selectbox("Top", [10, 20, 50, 100, 200], index=1)
    with c2:
        min_aantal = st.slider("Minimum aantal", 1, 50, 1)
    with c3:
        bt_vals = (
            df_schade_view["bus/tram"]
            .fillna("")
            .astype(str)
            .str.strip()
            .replace("", "(onbekend)")
            .unique()
            .tolist()
        )
        bt_vals = sorted(bt_vals)
        bus_tram = st.selectbox("Bus/Tram", ["Alles"] + bt_vals, index=0)
    with c4:
        type_vals = (
            df_schade_view["type"]
            .fillna("")
            .astype(str)
            .str.strip()
            .replace("", "(onbekend)")
            .unique()
            .tolist()
        )
        type_vals = sorted(type_vals)
        type_filter = st.selectbox("Type", ["Alles"] + type_vals, index=0)
    with c5:
        locatie_q = st.text_input("Zoek locatie", placeholder="bv. Gent, stelplaats, ...").strip().lower()

    tmp = df_schade_view.copy()

    tmp["Locatie"] = tmp["Locatie"].fillna("").astype(str).str.strip()
    tmp["bus/tram"] = tmp["bus/tram"].fillna("").astype(str).str.strip().replace("", "(onbekend)")
    tmp["type"] = tmp["type"].fillna("").astype(str).str.strip().replace("", "(onbekend)")
    tmp["voertuig"] = tmp["voertuig"].fillna("").astype(str).str.strip().replace("", "(onbekend)")
    tmp["teamcoach"] = tmp["teamcoach"].fillna("").astype(str).str.strip().replace("", "(onbekend)")

    tmp["Locatie"] = tmp["Locatie"].replace("", "(onbekend)")

    if bus_tram != "Alles":
        tmp = tmp[tmp["bus/tram"] == bus_tram].copy()

    if type_filter != "Alles":
        tmp = tmp[tmp["type"] == type_filter].copy()

    if locatie_q:
        tmp = tmp[tmp["Locatie"].str.lower().str.contains(re.escape(locatie_q), na=False)].copy()

    total_cases = len(tmp)
    unique_locaties = tmp["Locatie"].nunique(dropna=True)
    avg_per_loc = (total_cases / unique_locaties) if unique_locaties else 0.0

    top_loc = ""
    top_loc_count = 0
    if total_cases > 0:
        lc = tmp.groupby("Locatie").size().sort_values(ascending=False)
        if len(lc) > 0:
            top_loc = str(lc.index[0])
            top_loc_count = int(lc.iloc[0])

    k1, k2, k3, k4 = st.columns(4)
    k1.metric("Schadegevallen", f"{total_cases}")
    k2.metric("Unieke locaties", f"{unique_locaties}")
    k3.metric("Gemiddeld / locatie", f"{avg_per_loc:.2f}")
    k4.metric("Top locatie", f"{top_loc_count} — {top_loc}" if top_loc else "—")

    st.divider()

    st.markdown("### 📍 Hotspots: locaties met meeste schadegevallen")

    locaties_counts = (
        tmp.groupby("Locatie")
        .size()
        .reset_index(name="Aantal")
        .sort_values("Aantal", ascending=False)
    )
    locaties_counts = locaties_counts[locaties_counts["Aantal"] >= min_aantal].copy()

    top_table = locaties_counts.head(top_n)

    st.dataframe(
        top_table,
        use_container_width=True,
        hide_index=True,
        column_config={
            "Locatie": st.column_config.TextColumn("Locatie", width="large"),
            "Aantal": st.column_config.NumberColumn("Aantal", width="small"),
        },
    )

    if top_table.empty:
        st.caption("Geen locaties binnen deze filters.")
        st.stop()

    st.divider()

    st.markdown("### 📈 Trend & details voor gekozen locatie")

    locatie_options = top_table["Locatie"].tolist()
    gekozen_locatie = st.selectbox("Kies locatie", locatie_options, index=0)

    ldf = tmp[tmp["Locatie"] == gekozen_locatie].copy()

    ldf["Maand"] = ldf["Datum"].apply(_to_month)
    per_maand = (
        ldf[ldf["Maand"] != ""]
        .groupby("Maand")
        .size()
        .reset_index(name="Aantal")
        .sort_values("Maand")
    )

    cL, cR = st.columns([1.2, 1.0], gap="large")

    with cL:
        st.markdown("#### Schade per maand")
        if per_maand.empty:
            st.caption("Geen geldige datums om per maand te groeperen.")
        else:
            st.dataframe(per_maand, use_container_width=True, hide_index=True)
            st.bar_chart(per_maand.set_index("Maand")["Aantal"])

    with cR:
        st.markdown("#### Breakdown (top 10)")

        per_type = (
            ldf.groupby("type")
            .size()
            .reset_index(name="Aantal")
            .sort_values("Aantal", ascending=False)
            .head(10)
        )
        st.caption("Type")
        st.dataframe(per_type, use_container_width=True, hide_index=True)

        per_voertuig = (
            ldf.groupby("voertuig")
            .size()
            .reset_index(name="Aantal")
            .sort_values("Aantal", ascending=False)
            .head(10)
        )
        st.caption("Voertuig")
        st.dataframe(per_voertuig, use_container_width=True, hide_index=True)

        per_teamcoach = (
            ldf.groupby("teamcoach")
            .size()
            .reset_index(name="Aantal")
            .sort_values("Aantal", ascending=False)
            .head(10)
        )
        st.caption("Teamcoach")
        st.dataframe(per_teamcoach, use_container_width=True, hide_index=True)

elif current_page == "analyse":
    st.subheader("Analyse")

    if df_schade_view.empty:
        st.info("Geen schadegegevens beschikbaar voor deze selectie.")
        st.stop()

    def to_dt(v):
        return pd.to_datetime(v, dayfirst=True, errors="coerce")

    tmp = df_schade_view.copy()
    tmp["_dt"] = tmp["Datum"].apply(to_dt)
    tmp = tmp.dropna(subset=["_dt"])

    if tmp.empty:
        st.warning("Geen geldige datums gevonden om analyse te maken.")
        st.stop()

    for c in ["type", "Locatie", "voertuig", "teamcoach"]:
        if c in tmp.columns:
            tmp[c] = (
                tmp[c]
                .fillna("")
                .astype(str)
                .str.strip()
                .replace("", "(onbekend)")
            )

    st.markdown("## 📈 Evolutie doorheen de tijd")

    granularity = st.selectbox("Groeperen per", ["Maand", "Kwartaal"], index=0)

    if granularity == "Maand":
        tmp["Periode"] = tmp["_dt"].dt.to_period("M").astype(str)
    else:
        tmp["Periode"] = tmp["_dt"].dt.to_period("Q").astype(str)

    evolutie = (
        tmp.groupby("Periode")
        .size()
        .reset_index(name="Aantal schadegevallen")
        .sort_values("Periode")
    )

    c1, c2 = st.columns([1.1, 1.0])
    with c1:
        st.dataframe(evolutie, use_container_width=True, hide_index=True)
    with c2:
        st.bar_chart(evolutie.set_index("Periode")["Aantal schadegevallen"])

    st.divider()

    st.markdown("## 🧩 Verdeling per type")

    per_type = (
        tmp.groupby("type")
        .size()
        .reset_index(name="Aantal")
        .sort_values("Aantal", ascending=False)
    )
    total = per_type["Aantal"].sum()
    per_type["Aandeel (%)"] = (per_type["Aantal"] / total * 100).round(1)

    st.dataframe(
        per_type.head(10),
        use_container_width=True,
        hide_index=True,
        column_config={
            "Aandeel (%)": st.column_config.NumberColumn("Aandeel (%)", format="%.1f"),
        },
    )

    st.divider()

    st.markdown("## 🔥 Hotspot-combinaties")

    c1, c2 = st.columns(2)

    with c1:
        st.markdown("### Locatie × Type")
        loc_type = (
            tmp.groupby(["Locatie", "type"])
            .size()
            .reset_index(name="Aantal")
            .sort_values("Aantal", ascending=False)
            .head(10)
        )
        st.dataframe(loc_type, use_container_width=True, hide_index=True)

    with c2:
        st.markdown("### Voertuig × Type")
        veh_type = (
            tmp.groupby(["voertuig", "type"])
            .size()
            .reset_index(name="Aantal")
            .sort_values("Aantal", ascending=False)
            .head(10)
        )
        st.dataframe(veh_type, use_container_width=True, hide_index=True)

    st.divider()

    st.markdown("## 🚨 Recente signalen (laatste 6 maanden)")

    max_dt = tmp["_dt"].max()
    cutoff = max_dt - pd.DateOffset(months=6)

    recent = tmp[tmp["_dt"] >= cutoff].copy()

    if recent.empty:
        st.caption("Geen schadegevallen in de laatste 6 maanden.")
        st.stop()

    c1, c2, c3 = st.columns(3)

    with c1:
        st.markdown("### Locaties")
        recent_loc = (
            recent.groupby("Locatie")
            .size()
            .reset_index(name="Aantal")
            .sort_values("Aantal", ascending=False)
            .head(10)
        )
        st.dataframe(recent_loc, use_container_width=True, hide_index=True)

    with c2:
        st.markdown("### Voertuigen")
        recent_veh = (
            recent.groupby("voertuig")
            .size()
            .reset_index(name="Aantal")
            .sort_values("Aantal", ascending=False)
            .head(10)
        )
        st.dataframe(recent_veh, use_container_width=True, hide_index=True)

    with c3:
        st.markdown("### Types")
        recent_type = (
            recent.groupby("type")
            .size()
            .reset_index(name="Aantal")
            .sort_values("Aantal", ascending=False)
            .head(10)
        )
        st.dataframe(recent_type, use_container_width=True, hide_index=True)

    st.caption(
        f"Analyseperiode: {cutoff.strftime('%d-%m-%Y')} → {max_dt.strftime('%d-%m-%Y')}"
    )

elif current_page == "coaching":
    st.subheader("Coaching – Automatische selectie")

    min_schades = st.slider("Minimum aantal schades", 2, 20, 3)
    alleen_laatste_12m = st.checkbox(
        "Verwijder chauffeurs die > 1 jaar schadevrij zijn",
        value=True
    )

    def _to_dt(v):
        return pd.to_datetime(v, dayfirst=True, errors="coerce")

    def _mode_nonempty(s: pd.Series) -> str:
        s = s.dropna().astype(str).str.strip()
        s = s[(s != "") & (s.str.lower() != "(onbekend)") & (s.str.upper() != "#N/A")]
        if s.empty:
            return ""
        return s.value_counts().index[0]

    schade = df_schade_view.copy()
    if schade.empty:
        st.info("Geen schadegegevens in deze selectie.")
        st.stop()

    schade["personeelsnr"] = schade["personeelsnr"].apply(clean_id)
    schade["Datum_dt"] = schade["Datum"].apply(_to_dt)

    schade["Locatie"] = schade["Locatie"].fillna("").astype(str).str.strip()
    schade["type"] = schade["type"].fillna("").astype(str).str.strip()

    per_driver = (
        schade.groupby("personeelsnr", dropna=False)
        .agg(
            Schades=("personeelsnr", "size"),
            LaatsteDatum=("Datum_dt", "max"),
            TopLocatie=("Locatie", _mode_nonempty),
            TopType=("type", _mode_nonempty),
        )
        .reset_index()
    )

    per_driver["LaatsteDatum"] = pd.to_datetime(per_driver["LaatsteDatum"], errors="coerce")
    per_driver["LaatsteDatum_fmt"] = per_driver["LaatsteDatum"].dt.strftime("%d-%m-%Y").fillna("")

    planned_ids = set(
        df_coach_tab["nummer"].dropna().astype(str).apply(clean_id)
    ) if not df_coach_tab.empty else set()

    done_ids = set(
        df_coach_voltooid["nummer"].dropna().astype(str).apply(clean_id)
    ) if not df_coach_voltooid.empty else set()

    exclude_ids = planned_ids.union(done_ids)

    kandidaten = per_driver[
        (per_driver["Schades"] >= min_schades)
        & (~per_driver["personeelsnr"].isin(exclude_ids))
    ].copy()

    if alleen_laatste_12m:
        ref_dt = pd.to_datetime(schade["Datum_dt"], errors="coerce").max()
        if pd.notna(ref_dt):
            cutoff = ref_dt - pd.Timedelta(days=365)
            kandidaten = kandidaten[kandidaten["LaatsteDatum"] >= cutoff].copy()

    kandidaten = kandidaten.sort_values(
        ["Schades", "LaatsteDatum"],
        ascending=[False, False]
    )

    k1, k2, k3 = st.columns(3)
    k1.metric("Unieke chauffeurs met schade (jaar)", f"{per_driver['personeelsnr'].nunique()}")
    k2.metric(f"Kandidaten (≥{min_schades})", f"{len(kandidaten)}")
    k3.metric("Uitgesloten (gepland of voltooid)", f"{len(exclude_ids)}")

    st.divider()

    st.markdown("### 🎯 Chauffeurs te coachen")

    if kandidaten.empty:
        st.success("Geen chauffeurs die voldoen aan de criteria 🎉")
        st.stop()

    show_cols = [
        "personeelsnr",
        "Schades",
        "LaatsteDatum_fmt",
        "TopLocatie",
        "TopType",
    ]

    st.dataframe(
        kandidaten[show_cols],
        use_container_width=True,
        hide_index=True,
        column_config={
            "personeelsnr": st.column_config.TextColumn("P-nr", width="small"),
            "Schades": st.column_config.NumberColumn("Schades", width="small"),
            "LaatsteDatum_fmt": st.column_config.TextColumn("Laatste schade", width="small"),
            "TopLocatie": st.column_config.TextColumn("Top locatie", width="medium"),
            "TopType": st.column_config.TextColumn("Top type", width="medium"),
        },
    )

    csv_bytes = kandidaten[show_cols].to_csv(index=False).encode("utf-8")
    st.download_button(
        "⬇️ Download coaching kandidaten (CSV)",
        data=csv_bytes,
        file_name="coaching_kandidaten.csv",
        mime="text/csv",
        use_container_width=True,
    )
